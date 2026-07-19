"""
run_predictions.py
----------------------
Standalone script to run medical predictions on an excel sample file.

Usage:
    python run_predictions.py --input data/sample.xlsx --output results/predictions.xlsx

Features:
- Processes data in batches of 10 rows (configurable with --batch-size)
- Saves a checkpoint .xlsx after every batch
- Saves full logs to a .log file alongside the output
- Prints a live progress bar to the console
- Resumes from the last checkpoint if interrupted (skips already-processed VisitServiceIDs)
"""

import argparse
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tqdm import tqdm

# Allow importing predictions.py from the same directory or src/
sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent / "src"))

from predictions_openrouter import make_preds


# ─────────────────────────────────────────────
# Logging setup
# ─────────────────────────────────────────────

def setup_logging(log_path: Path):
    root = logging.getLogger()
    root.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # File handler – captures everything
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    root.addHandler(fh)

    # Console handler – INFO and above only
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    root.addHandler(ch)

    return logging.getLogger("run_predictions_csv")


# ─────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
APPROVED_FILL = PatternFill("solid", start_color="C6EFCE")
REJECTED_FILL = PatternFill("solid", start_color="FFC7CE")
FAILED_FILL   = PatternFill("solid", start_color="FFEB9C")
ROW_FONT      = Font(name="Arial", size=10)

COL_WIDTHS = {
    "VisitID": 14,
    "VisitServiceID": 18,
    "Service_Name": 30,
    "Medical_Prediction": 20,
    "Reason/Recommendation": 60,
    "Diagnose": 30,
    "Chief_Complaint": 30,
    "ProblemNote": 40,
    "Symptoms": 40,
}


def style_sheet(ws):
    """Apply header formatting and column widths to a worksheet."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, cell in enumerate(ws[1], start=1):
        col_letter = cell.column_letter
        col_name = cell.value or ""
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_name, 18)

    for row in ws.iter_rows(min_row=2):
        pred_val = None
        for cell in row:
            cell.font = ROW_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if ws.cell(row=1, column=cell.column).value == "Medical_Prediction":
                pred_val = cell.value

        if pred_val == "Approved":
            fill = APPROVED_FILL
        elif pred_val == "Rejected":
            fill = REJECTED_FILL
        elif pred_val and "Failed" in str(pred_val):
            fill = FAILED_FILL
        else:
            fill = None

        if fill:
            for cell in row:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_xlsx(df: pd.DataFrame, path: Path):
    """Write DataFrame to xlsx with formatting."""
    df.to_excel(path, index=False, engine="openpyxl")
    wb = load_workbook(path)
    style_sheet(wb.active)
    wb.save(path)


# ─────────────────────────────────────────────
# Checkpoint helpers
# ─────────────────────────────────────────────

def load_checkpoint(checkpoint_path: Path) -> pd.DataFrame:
    if checkpoint_path.exists():
        return pd.read_excel(checkpoint_path)
    return pd.DataFrame()


def save_checkpoint(all_results: pd.DataFrame, checkpoint_path: Path):
    save_xlsx(all_results, checkpoint_path)


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Run medical predictions on a CSV sample.")
    p.add_argument("--input",      required=True,  help="Path to input CSV file")
    p.add_argument("--output",     required=True,  help="Path to final output .xlsx file")
    p.add_argument("--batch-size", type=int, default=10, help="Rows per batch (default: 10)")
    p.add_argument("--model",      default="deepseek/deepseek-chat-v3.1",
                   help="OpenRouter model identifier (default: deepseek/deepseek-chat-v3.1)")
    return p.parse_args()


def main():
    args = parse_args()

    input_path      = Path(args.input)
    output_path     = Path(args.output)
    batch_size      = args.batch_size
    model_name      = args.model

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Derive log and checkpoint paths from output
    timestamp       = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path        = output_path.with_suffix("").with_name(output_path.stem + f"_{timestamp}.log")
    checkpoint_path = output_path.with_suffix("").with_name(output_path.stem + "_checkpoint.xlsx")

    logger = setup_logging(log_path)
    logger.info(f"Input:       {input_path}")
    logger.info(f"Output:      {output_path}")
    logger.info(f"Checkpoint:  {checkpoint_path}")
    logger.info(f"Log:         {log_path}")
    logger.info(f"Batch size:  {batch_size}")
    logger.info(f"Model:       {model_name}")

    # ── Load data ──────────────────────────────
    if not input_path.exists():
        logger.error(f"Input file not found: {input_path}")
        sys.exit(1)

    # Detect file type
    if input_path.suffix.lower() in [".xlsx", ".xls"]:
        try:
            df = pd.read_excel(input_path)
            logger.info(f"Loaded {len(df)} rows from Excel file: {input_path}")
        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            sys.exit(1)

    elif input_path.suffix.lower() == ".csv":
        for encoding in ("utf-8", "cp1252", "latin-1"):
            try:
                df = pd.read_csv(input_path, encoding=encoding)
                logger.info(f"Loaded {len(df)} rows from CSV: {input_path} (encoding: {encoding})")
                break
            except UnicodeDecodeError:
                logger.warning(f"Encoding {encoding} failed, trying next…")
        else:
            logger.error("Could not read CSV with utf-8, cp1252, or latin-1 encodings")
            sys.exit(1)

    else:
        logger.error(f"Unsupported file format: {input_path.suffix}")
        sys.exit(1)

    # ── Resume from checkpoint ─────────────────
    checkpoint_df = load_checkpoint(checkpoint_path)
    already_done_ids = set()
    all_results = pd.DataFrame()

    if not checkpoint_df.empty and "VisitServiceID" in checkpoint_df.columns:
        already_done_ids = set(checkpoint_df["VisitServiceID"].astype(str))
        all_results = checkpoint_df.copy()
        logger.info(f"Resuming: {len(already_done_ids)} VisitServiceIDs already processed")

    df["VisitServiceID"] = df["VisitServiceID"].astype(str)
    remaining_df = df[~df["VisitServiceID"].isin(already_done_ids)].copy()
    logger.info(f"Rows remaining to process: {len(remaining_df)}")

    if remaining_df.empty:
        logger.info("Nothing to do – all rows already processed.")
        save_xlsx(all_results, output_path)
        logger.info(f"Final output saved to {output_path}")
        return

    # ── Batch loop ─────────────────────────────
    total_rows    = len(remaining_df)
    total_batches = (total_rows + batch_size - 1) // batch_size

    cumulative_cost   = 0.0
    cumulative_input  = 0
    cumulative_output = 0

    with tqdm(total=total_rows, desc="Processing rows", unit="row") as pbar:
        for batch_idx in range(total_batches):
            start = batch_idx * batch_size
            end   = min(start + batch_size, total_rows)
            batch = remaining_df.iloc[start:end].copy()

            logger.info(
                f"[Batch {batch_idx + 1}/{total_batches}] "
                f"Rows {start}–{end - 1} | {len(batch)} rows"
            )

            try:
                # Monkey-patch the model name into the predictions module
                import predictions_openrouter as pred_module
                _orig = pred_module.dev_response

                def _patched_dev_response(info, services, model=model_name):
                    return _orig(info, services, model=model)

                pred_module.dev_response = _patched_dev_response

                result_df, metrics = make_preds(batch)

                pred_module.dev_response = _orig  # restore

            except Exception as e:
                logger.error(f"[Batch {batch_idx + 1}] Failed: {e}", exc_info=True)
                pbar.update(len(batch))
                continue

            # Accumulate cost
            cumulative_cost   += metrics["total_cost"]
            cumulative_input  += metrics["total_input_tokens"]
            cumulative_output += metrics["total_output_tokens"]

            logger.info(
                f"[Batch {batch_idx + 1}] Cost: ${metrics['total_cost']:.4f} | "
                f"Tokens in/out: {metrics['total_input_tokens']}/{metrics['total_output_tokens']}"
            )

            # Merge results
            all_results = pd.concat([all_results, result_df], ignore_index=True)

            # Save checkpoint
            save_checkpoint(all_results, checkpoint_path)
            logger.info(
                f"[Batch {batch_idx + 1}] Checkpoint saved → {checkpoint_path} "
                f"(total rows so far: {len(all_results)})"
            )

            pbar.update(len(batch))
            pbar.set_postfix(
                batches_done=f"{batch_idx + 1}/{total_batches}",
                cost=f"${cumulative_cost:.4f}",
            )

    # ── Final output ───────────────────────────
    logger.info("All batches complete. Writing final output…")
    save_xlsx(all_results, output_path)
    logger.info(f"Final output saved → {output_path}")

    logger.info("=" * 50)
    logger.info("FINAL COST SUMMARY")
    logger.info(f"  Total input tokens:  {cumulative_input:,}")
    logger.info(f"  Total output tokens: {cumulative_output:,}")
    logger.info(f"  Total tokens:        {cumulative_input + cumulative_output:,}")
    logger.info(f"  Total cost:          ${cumulative_cost:.4f}")
    logger.info("=" * 50)

    print(f"\n✅ Done! Results saved to: {output_path}")
    print(f"📋 Log saved to:          {log_path}")
    print(f"💰 Total cost:            ${cumulative_cost:.4f}")


if __name__ == "__main__":
    main()